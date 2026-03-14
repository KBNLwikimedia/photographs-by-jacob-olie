"""
Extract full metadata from the Memorix Mediabank API for all records and add
them as new columns to the Excel workbook.

Reads identifiers from column D (Archief Amsterdam URL), queries the Memorix
API, and writes 13 metadata columns (G onwards) to jacob_olie_sources.xlsx.

Multi-value fields are formatted as semicolon-separated double-quoted strings,
e.g. "Frankendael"; "Munttoren".
"""

import re
import time
import requests
import openpyxl

API_URL = "https://webservices.memorix.nl/mediabank/media"
API_KEY = "eb37e65a-eb47-11e9-b95c-60f81db16c0e"
USER_AGENT = "JacobOlieSourceExtractor/1.0 (https://github.com/KBNLresearch; photographs-by-jacob-olie project) Python/requests"
DELAY = 0.5

# Ordered list of (api_field, label) for the columns we want to extract
METADATA_FIELDS = [
    ("dc_title", "Titel"),
    ("dc_description", "Beschrijving"),
    ("dc_date", "Datering"),
    ("sk_documenttype", "Documenttype"),
    ("sk_vervaardiger", "Vervaardiger"),
    ("dc_provenance", "Collectie"),
    ("geografische_aanduiding", "Geografische aanduiding"),
    ("sk_gebouw", "Gebouw"),
    ("dc_source", "Inventarissen"),
    ("identifier", "Afbeeldingsbestand"),
    ("sr_rechthebbende", "Rechthebbende"),
    ("sr_leveringsvoorwaarden", "Gebruiksvoorwaarden"),
    ("quality", "Kwaliteit"),
]

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def format_value(value):
    """Format a metadata value for Excel output.

    Handles strings, lists, and nested structures (like Geografische aanduiding).
    Multi-value fields are returned as semicolon-separated double-quoted strings.

    Args:
        value: A string, list of strings, or nested list of dicts from the API.

    Returns:
        A formatted string suitable for an Excel cell.
    """
    if value is None:
        return ""

    if isinstance(value, str):
        # Strip HTML tags (e.g. <a href=...>...</a> in Inventarissen)
        clean = re.sub(r'<[^>]+>', '', value).strip()
        return clean

    if isinstance(value, list):
        # Check if it's a nested geo structure: list of lists of dicts
        if value and isinstance(value[0], list):
            # Geografische aanduiding: flatten nested dicts
            parts = []
            for group in value:
                if isinstance(group, list):
                    for item in group:
                        if isinstance(item, dict):
                            label = item.get("label", "")
                            val = item.get("value", "")
                            if val:
                                parts.append(f"{label}: {val}")
            if parts:
                return "; ".join(f'"{p}"' for p in parts)
            return ""

        # Simple list of strings
        if len(value) == 1:
            return format_value(value[0])
        return "; ".join(f'"{format_value(v)}"' for v in value)

    return str(value)


def extract_metadata(record):
    """Extract all metadata fields from a Memorix API record.

    Args:
        record: A single record dict from the API's ``media`` array.

    Returns:
        A dict mapping API field names to formatted string values.
    """
    result = {}
    for md in record.get("metadata", []):
        field = md.get("field", "")
        value = md.get("value", "")
        result[field] = format_value(value)
    return result


def main():
    """Main entry point: read Excel, query API, write metadata columns."""
    xlsx_path = "jacob_olie_sources.xlsx"
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Write headers for metadata columns starting at column G (7)
    col_start = 7
    for i, (api_field, label) in enumerate(METADATA_FIELDS):
        ws.cell(row=1, column=col_start + i, value=f"{label} ({api_field})")

    # Collect rows with Archief Amsterdam URL (column D)
    rows_to_process = []
    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=4).value
        if url and url.strip():
            m = re.search(r'[?&]q=([^&]+)', url)
            if m:
                rows_to_process.append((row_idx, m.group(1)))

    total = len(rows_to_process)
    print(f"Found {total} rows to process")

    found = 0
    not_found = 0
    errors = 0

    for i, (row_idx, identifier) in enumerate(rows_to_process):
        try:
            params = {"q": identifier, "rows": 1, "page": 1, "apiKey": API_KEY}
            resp = session.get(API_URL, params=params)
            resp.raise_for_status()
            data = resp.json()
            media = data.get("media", [])

            if media:
                metadata = extract_metadata(media[0])
                for j, (api_field, _) in enumerate(METADATA_FIELDS):
                    val = metadata.get(api_field, "")
                    ws.cell(row=row_idx, column=col_start + j).value = val
                found += 1
            else:
                not_found += 1
        except Exception as e:
            ws.cell(row=row_idx, column=col_start).value = f"ERROR: {e}"
            errors += 1

        if (i + 1) % 100 == 0 or (i + 1) == total:
            print(f"  {i + 1}/{total} — found: {found}, not found: {not_found}, errors: {errors}")

        time.sleep(DELAY)

    # Auto-size new columns
    for i in range(len(METADATA_FIELDS)):
        col_letter = openpyxl.utils.get_column_letter(col_start + i)
        ws.column_dimensions[col_letter].width = 35

    wb.save(xlsx_path)
    print(f"\nDone! Updated {xlsx_path}")
    print(f"  Metadata found: {found}")
    print(f"  Not found: {not_found}")
    print(f"  Errors: {errors}")


if __name__ == "__main__":
    main()

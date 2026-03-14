"""
Resolve Stadsarchief Amsterdam identifiers to Beeldbank detail page URLs.

This script reads the Excel workbook produced by ``extract_sources.py``, looks
at the "Archief Amsterdam URL" column (D), extracts the record identifier from
the query string, and queries the Memorix Mediabank API to obtain the internal
UUID for each record. It then constructs the canonical detail page URL and
writes it into a new column (E) in the same workbook.

The Stadsarchief Amsterdam Beeldbank (https://archief.amsterdam/beeldbank/) is
powered by Vitec Memorix. The public Memorix Mediabank API is at
``https://webservices.memorix.nl/mediabank/media`` and requires an API key,
which is embedded in the Beeldbank's HTML page source (public, not secret).

Requests are throttled to one every 0.5 seconds to be respectful to the server.

Prerequisites:
    pip install requests openpyxl

Usage:
    python add_detail_urls.py

Input/Output:
    jacob_olie_sources.xlsx — the workbook is read and updated in place.
    Column D ("Archief Amsterdam URL") must already be populated (e.g. by a
    prior URL-transformation step). The script adds column E
    ("Archief Amsterdam Detail URL").

Adapting for other collections:
    See MANUAL.md for instructions on how to reuse this workflow for other
    Stadsarchief Amsterdam or Memorix-based collections.
"""

import re
import time
import requests
import openpyxl

API_URL = "https://webservices.memorix.nl/mediabank/media"
API_KEY = "eb37e65a-eb47-11e9-b95c-60f81db16c0e"
DETAIL_BASE = "https://archief.amsterdam/beeldbank/detail/"
USER_AGENT = "JacobOlieSourceExtractor/1.0 (https://github.com/KBNLresearch; photographs-by-jacob-olie project) Python/requests"
DELAY = 0.5  # seconds between API calls (be gentle)

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def extract_identifier(archief_url):
    """Extract the record identifier from an Archief Amsterdam gallery URL.

    The Beeldbank gallery URLs contain the identifier in the ``q`` query
    parameter, e.g.::

        https://archief.amsterdam/beeldbank/?mode=gallery&view=horizontal&q=010019000001&rows=1&page=1

    Args:
        archief_url: A full Archief Amsterdam Beeldbank gallery URL.

    Returns:
        The identifier string (e.g. ``"010019000001"``), or None if the
        ``q`` parameter is not found.
    """
    m = re.search(r'[?&]q=([^&]+)', archief_url)
    return m.group(1) if m else None


def lookup_record_id(identifier):
    """Look up a record's UUID via the Memorix Mediabank API.

    Sends a search query for the given identifier and returns the UUID of
    the first matching record. The UUID is used to construct the canonical
    detail page URL on the Beeldbank website.

    Args:
        identifier: A Stadsarchief Amsterdam record identifier, e.g.
                     ``"010019000001"`` or ``"10019A001542"`` or ``"BMAB00003000001"``.

    Returns:
        The record UUID as a string (e.g.
        ``"bf2bc41b-9441-9049-1f28-5012c8617cc3"``), or None if no matching
        record was found.

    Raises:
        requests.HTTPError: If the API returns a non-2xx status code.
    """
    params = {
        "q": identifier,
        "rows": 1,
        "page": 1,
        "apiKey": API_KEY,
    }
    resp = session.get(API_URL, params=params)
    resp.raise_for_status()
    data = resp.json()
    media = data.get("media", [])
    if media:
        return media[0].get("id")
    return None


def main():
    """Main entry point: read Excel, query Memorix API, write detail URLs."""
    xlsx_path = "jacob_olie_sources.xlsx"
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Add header for new column
    ws.cell(row=1, column=5, value="Archief Amsterdam Detail URL")

    # Collect rows that have an Archief Amsterdam URL (column D)
    rows_to_process = []
    for row_idx in range(2, ws.max_row + 1):
        archief_url = ws.cell(row=row_idx, column=4).value
        if archief_url and archief_url.strip():
            rows_to_process.append(row_idx)

    total = len(rows_to_process)
    print(f"Found {total} rows with Archief Amsterdam URLs to look up")

    found = 0
    not_found = 0
    errors = 0

    for i, row_idx in enumerate(rows_to_process):
        archief_url = ws.cell(row=row_idx, column=4).value
        identifier = extract_identifier(archief_url)

        if not identifier:
            ws.cell(row=row_idx, column=5).value = "ERROR: no identifier"
            errors += 1
            continue

        try:
            record_id = lookup_record_id(identifier)
            if record_id:
                detail_url = DETAIL_BASE + record_id
                ws.cell(row=row_idx, column=5).value = detail_url
                found += 1
            else:
                ws.cell(row=row_idx, column=5).value = "NOT FOUND"
                not_found += 1
        except Exception as e:
            ws.cell(row=row_idx, column=5).value = f"ERROR: {e}"
            errors += 1

        # Progress update every 100 rows
        if (i + 1) % 100 == 0 or (i + 1) == total:
            print(f"  {i + 1}/{total} — found: {found}, not found: {not_found}, errors: {errors}")

        # Rate limiting
        time.sleep(DELAY)

    ws.column_dimensions['E'].width = 80
    wb.save(xlsx_path)
    print(f"\nDone! Updated {xlsx_path}")
    print(f"  Detail URLs found: {found}")
    print(f"  Not found: {not_found}")
    print(f"  Errors: {errors}")


if __name__ == "__main__":
    main()
